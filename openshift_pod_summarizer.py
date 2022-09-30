#!/usr/bin/python3

import sys
import copy
import json
import yaml
import pprint
import openpyxl
import argparse
import subprocess
from urllib.parse import urlparse

alldata = {}

header_labels = [
    'ns',
    'pod_name',
    'description',
    'url',
    'custom_resources',
    'how_to_install',
    'num_of_pods',
    'owner_kind',
    'owner_name',
    'affinity',
    'dnsPolicy',
    'enableServiceLinks',
    'hostNetwork',
    'hostPID',
    # 'nodeName',
    # 'role',
    'nodeSelector',
    'preemptionPolicy',
    'priority',
    'priorityClassName',
    'restartPolicy',
    'schedulerName',
    'serviceAccount',
    'serviceAccountName',
    'pod_securityContext',
    'tolerations',
    'terminationGracePeriodSeconds',
    'qosClass',
    'container_name',
    'initContainer',
    'container_image',
    'container_imagePullPolicy',
    'container_resources',
    'container_securityContext'
]

header2column = {}
for i, label in enumerate(header_labels, 1):
    # openxl column is 1-origin
    header2column[label] = i

def load_desc(path):
    desc = {}
    desc_hash = {}
    with open(path, 'r') as f:
        desc = yaml.load(f, Loader=yaml.FullLoader)
    for item in desc['descriptions']:
        desc_hash[(item['ns'], item['name'])] = {'desc':item.get('desc', ''), 'url':item.get('url', ''), 'crd':item.get('crd', ''), 'how_to_install':item.get('install', '')}
    return desc_hash

def load_nodes():
    masters = []
    workers = []
    for item in alldata['Node']:
        hostname = item['metadata']['name']
        labels = item['metadata']['labels']
        if 'node-role.kubernetes.io/master' in labels:
            masters.append(hostname)
        if 'node-role.kubernetes.io/worker' in labels:
            workers.append(hostname)
    return masters, workers

# def load_offline_data(json_path):
#     json_data = {}
#     if json_path:
#         with open(json_path, 'r') as f:
#             print('** json from {}'.format(json_path))
#             json_data = json.load(f)
#     else:
#         print('** json from `kubectl get -A pod,node,replicaset,statefulset,deployment,catalogsource -o json`')
#         output = subprocess.run('kubectl get -A pod,node,replicaset,statefulset,deployment,catalogsource -o json'.split(), capture_output=True)
#         json_data = json.loads(output.stdout)
# 
#     for item in json_data['items']:
#         kind = item['kind']
#         if not json_data.get(kind):
#             json_data[kind] = []
#         json_data[kind].append(item)
# 
#     return json_data

def load_offline_data(json_files):
    json_data = {}
    json_data['raw'] = []
    if json_files:
        for file in json_files:
            with open(file, 'r') as f:
                print('** json from {}'.format(file))
                json_data['raw'].append(json.load(f))

    else:
        print('** json from `kubectl get -A pod,node,replicaset,statefulset,deployment,catalogsource -o json`')
        output = subprocess.run('kubectl get -A pod,node,replicaset,statefulset,deployment,catalogsource -o json'.split(), capture_output=True)
        json_data['raw'].append(json.loads(output.stdout))

    for data in json_data['raw']:
        for item in data['items']:
            kind = item['kind']
            if not json_data.get(kind):
                json_data[kind] = []
            json_data[kind].append(item)

    json_data['Pod'].sort(key=lambda x: (x['metadata']['namespace'], x['metadata']['name']))

    return json_data

def hostname2role(hostname):
    role = ''
    if hostname in masters:
        role = role + 'master'
    if hostname in workers:
        if role != '':
            role = role + '/'
        role = role + 'worker'
    return role

def print_nodes(masters, workers):
    print('# masters:')
    pprint.pprint(masters)
    print('# workers:')
    pprint.pprint(workers)

def write_row(sheet, array, start_row, start_col, color):
    for x, cell in enumerate(array):
        sheet.cell(row=start_row, column=start_col + x, value=array[x])
        sheet.cell(row=start_row, column=start_col + x).fill = color

def dict2json(obj):
    return json.dumps(obj, indent=2)

def dict2yaml(obj):
    return yaml.dump(obj, Dumper=yaml.Dumper).rstrip()

def ns_pod_key(ns, pod):
    return '{}__{}'.format(ns, pod)

def normalize_pod_name(name, owner_kind, node_name):
    array = name.split('-')
    if owner_kind == 'DaemonSet' or owner_kind == 'CatalogSource':
        # array[-1] = 'X' * len(array[-1])
        array[-1] = 'X' * 5
        return '-'.join(array)
    if owner_kind == 'ReplicaSet' or owner_kind == 'Job':
        # array[-2] = 'X' * len(array[-2])
        array[-2] = 'X' * 10
        # array[-1] = 'X' * len(array[-1])
        array[-1] = 'X' * 5
        return '-'.join(array)
    if owner_kind == 'StatefulSet':
        return name[0:-1] + 'X'
    if owner_kind == 'Node':
        return name.replace(node_name, 'HOSTNAME')
    if owner_kind == 'ConfigMap':
        array = name.replace(node_name, 'HOSTNAME').split('-')
        array[-2] = 'X'
        return '-'.join(array)
    if name.startswith('etcd-guard') or name.startswith('kube-apiserver-guard') or name.startswith('kube-controller-manager-guard') or name.startswith('openshift-kube-scheduler-guard'):
        return name.replace(node_name, 'HOSTNAME')
    return name

def normalize_owner_name(owner_kind, owner_name):
    print('  XXX normalize_owner_name():owner_kind={}, owner_name={}'.format(owner_kind, owner_name))
    array = owner_name.split('-')
    if owner_kind == 'ReplicaSet' or owner_kind == 'Job':
        # array[-1] = 'X' * len(array[-1])
        array[-1] = 'X' * 10
        return '-'.join(array)
    if owner_kind == 'Node':
        return 'HOSTNAME'
    return owner_name

def find_resource_json(kind, ns, name):
    if alldata.get(kind):
        for item in alldata[kind]:
            if item['metadata']['namespace'] == ns and item['metadata']['name'] == name:
                return item
    return None
    
def normalize_owner_kind(owner_kind, owner_name, ns):
    print('  XXX normalize_owner_kind():owner_kind={}, owner_name={}, ns={}'.format(owner_kind, owner_name, ns))
    if owner_kind == 'ReplicaSet' or owner_kind == 'Job':
        json_data = find_resource_json(owner_kind, ns, owner_name)
        refs = json_data['metadata'].get('ownerReferences')
        if refs:
            owner_owner_kind = refs[0]['kind']
            return '{} ({})'.format(owner_kind, owner_owner_kind)
    return owner_kind

def get_number_of_pods(selector, owner_kind, owner_name, pod, ns):
    print('  XXX get_number_of_pods():selector={}, owner_kind={}, owner_name={}, pod={}, ns={}'.format(selector, owner_kind, owner_name, pod, ns))
    if owner_kind == 'DaemonSet':
        print('  ### selector:{}'.format(selector))
        # if 'node-role.kubernetes.io/master' in selector:
        if selector.get('node-role.kubernetes.io/master') == '':
            return '# of masters'
        # if 'node-role.kubernetes.io/worker' in selector:
        if selector.get('node-role.kubernetes.io/worker') == '':
            return '# of workers'
        if selector.get('kubernetes.io/os') == 'linux' or selector.get('beta.kubernetes.io/os') == 'linux':
            return '# of linux nodes'
        return 'unknown'

    if owner_kind == 'Node' and ns == 'openshift-etcd':
        return 'Static Pod on masters'
    if owner_kind == 'ConfigMap' and ns == 'openshift-etcd':
        return 'Pod for maintaining Static Pod'
    if owner_kind == 'Node' and ns == 'openshift-kube-apiserver':
        return 'Static Pod on masters'
    if owner_kind == 'ConfigMap' and ns == 'openshift-kube-apiserver':
        return 'Pod for maintaining Static Pod'
    if owner_kind == 'Node' and ns == 'openshift-kube-controller-manager':
        return 'Static Pod on masters'
    if owner_kind == 'ConfigMap' and ns == 'openshift-kube-controller-manager':
        return 'Pod for maintaining Static Pod'
    if owner_kind == 'Node' and ns == 'openshift-kube-scheduler':
        return 'Static Pod on masters'
    if owner_kind == 'ConfigMap' and ns == 'openshift-kube-scheduler':
        return 'Pod for maintaining Static Pod'
    if owner_kind == 'Node' and (ns == 'openshift-kni-infra' or ns == 'openshift-nutanix-infra' or ns == 'openshift-openstack-infra' or ns == 'openshift-ovirt-infra' or ns == 'openshift-vsphere-infra'):
        if pod.startswith('haproxy'):
            return 'Static Pod on masters'
        else:
            return 'Static Pod on masters and workers'

    if owner_kind == 'CatalogSource' or owner_kind == 'Job':
        return ''

    json_data = find_resource_json(owner_kind, ns, owner_name)
    replicas = json_data['spec'].get('replicas')
    return 'replicas={}'.format(replicas)

def xls_input_cell_by_key(sheet, row, key, value):
    sheet.cell(row=row, column=header2column[key], value=value)
    sheet.cell(row=row, column=header2column[key]).alignment = openpyxl.styles.Alignment(vertical='center')
    sheet.cell(row=row, column=header2column[key]).font = openpyxl.styles.fonts.Font(name='Source Code Pro Medium')

def make_container_cells(sheet, current_row, ctr, is_init_container):
    xls_input_cell_by_key(sheet, current_row, 'container_name', ctr.get('name', ''))
    xls_input_cell_by_key(sheet, current_row, 'initContainer', str(is_init_container).lower() if is_init_container else '')
    xls_input_cell_by_key(sheet, current_row, 'container_image', ctr.get('image', ''))
    xls_input_cell_by_key(sheet, current_row, 'container_imagePullPolicy', ctr.get('imagePullPolicy', ''))
    xls_input_cell_by_key(sheet, current_row, 'container_resources', dict2yaml(ctr.get('resources', '')))
    xls_input_cell_by_key(sheet, current_row, 'container_securityContext', dict2yaml(ctr.get('securityContext', '')))

def get_desc_info(desc_hash, ns, pod_name, key):
    desc = desc_hash.get((ns, pod_name), None)
    if desc:
        return desc.get(key, '')
    return ''

def build_crd_str(desc, ns, pod_name):
    all_crd_str = ''
    crds = get_desc_info(desc, ns, pod_name, 'crd')
    # print('  => !!! crds={}'.format(crds))
    for crd in crds:
        kind = crd.get('kind')
        apiversion = crd.get('apiversion')

        if kind == None or kind == '':
            continue
        if apiversion == None or apiversion == '':
            apiversion = None
        else:
            apiversion = '(' + apiversion + ')'
        crd_str = kind + (' ' + apiversion) if apiversion else ''

        if all_crd_str == '':
            all_crd_str = crd_str
        else:
            all_crd_str = all_crd_str + ',\n' + crd_str
    # print('  => !!! all_crd_str={}'.format(all_crd_str))
    return all_crd_str

def set_cell_hyperlink(cell, url):
    if url == None or url == '':
        return

    parse_result = urlparse(url)

    if parse_result.path == '/':
        cell.value = url
        cell.hyperlink = url
        return

    path = parse_result.path[1:]
    if path[-1] == '/':
        path = path.rstrip()
    
    dirs = path.split('/')
    if len(dirs) >= 2:
        path = '/'.join(dirs[:2])

    cell.value = path
    cell.hyperlink = url

def set_cell_wrap_text(cell):
    new_alignment = copy.copy(cell.alignment)
    new_alignment.wrapText = True
    cell.alignment = new_alignment

def main(args):
    desc = load_desc(args.description_yaml)
    global alldata
    alldata = load_offline_data(args.offline)
    masters, workers = load_nodes()
    print_nodes(masters, workers)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Pods'
    current_row = 1
    fill_header = openpyxl.styles.PatternFill(patternType='solid', fgColor='D9EAD3')
    # sheet.freeze_panes = 'A2'
    sheet.freeze_panes = 'C2'

    pod_exists = {}

    write_row(sheet, header_labels, 1, 1, fill_header)
    current_row = current_row + 1

    for item in alldata['Pod']:
        md = item['metadata']
        spec = item['spec']
        status = item['status']
        row = []

        # if status['phase'] != 'R{}'.format(spec.get('affinity', '')))unning':
        #     continue

        refs = md.get('ownerReferences')
        # rename 'pod_name' column
        pod_name = normalize_pod_name(md['name'], refs[0]['kind'] if refs else '', spec['nodeName'])

        if pod_exists.get((md['namespace'], pod_name)):
            print('  => SKIP')
            continue
        pod_exists[(md['namespace'], pod_name)] = True

        print('* ns:{}, pod_name:{}, phase:{}'.format(md['namespace'], md['name'], status['phase']))
        xls_input_cell_by_key(sheet, current_row, 'ns', md['namespace'])
        xls_input_cell_by_key(sheet, current_row, 'pod_name', pod_name)

        xls_input_cell_by_key(sheet, current_row, 'description', get_desc_info(desc, md['namespace'], pod_name, 'desc'))
        set_cell_wrap_text(sheet.cell(row=current_row, column=header2column['description']))

        xls_input_cell_by_key(sheet, current_row, 'url', get_desc_info(desc, md['namespace'], pod_name, 'url'))
        set_cell_wrap_text(sheet.cell(row=current_row, column=header2column['url']))
        set_cell_hyperlink(sheet.cell(row=current_row, column=header2column['url']), get_desc_info(desc, md['namespace'], pod_name, 'url'))

        xls_input_cell_by_key(sheet, current_row, 'custom_resources', build_crd_str(desc, md['namespace'], pod_name))

        xls_input_cell_by_key(sheet, current_row, 'how_to_install', get_desc_info(desc, md['namespace'], pod_name, 'how_to_install'))
        set_cell_wrap_text(sheet.cell(row=current_row, column=header2column['how_to_install']))

        if refs:
            ref = refs[0]
            if len(refs) > 1:
                print('!! more than one ownerReferences !!')
                sys.exit(1)

            # 'number of pods' column
            xls_input_cell_by_key(sheet, current_row, 'num_of_pods', get_number_of_pods(spec.get('nodeSelector', ''), ref['kind'], ref['name'], pod_name, md['namespace']))
            # xls_input_cell_by_key(
            #     sheet,
            #     current_row,
            #     'num_of_pods',
            #     get_number_of_pods(
            #         spec.get('nodeSelector', ''),
            #         ref['kind'],
            #         normalize_owner_name(ref['kind'], ref['name']),
            #         pod_name, md['namespace']
            #     )
            # )

            xls_input_cell_by_key(sheet, current_row, 'owner_kind', normalize_owner_kind(ref['kind'], ref['name'], md['namespace']))
            xls_input_cell_by_key(sheet, current_row, 'owner_name', normalize_owner_name(ref['kind'], ref['name']))

        print('  affinity:{}'.format(spec.get('affinity', '')))
        xls_input_cell_by_key(sheet, current_row, 'affinity', dict2yaml(spec.get('affinity', '')))
        print('  dnsPolicy:{}, hostNetwork:{}, hostPID:{}'.format(spec.get('dnsPolicy', ''), spec.get('hostNetwork', ''), spec.get('hostPID', '')))
        xls_input_cell_by_key(sheet, current_row, 'dnsPolicy', spec.get('dnsPolicy', ''))
        xls_input_cell_by_key(sheet, current_row, 'enableServiceLinks', str(spec.get('enableServiceLinks', '')).lower())
        xls_input_cell_by_key(sheet, current_row, 'hostNetwork', str(spec.get('hostNetwork', '')).lower())
        xls_input_cell_by_key(sheet, current_row, 'hostPID', str(spec.get('hostPID', '')).lower())
        # print('  nodeName:{}, role:{}, nodeSelector:{}'.format(spec.get('nodeName', ''), hostname2role(spec.get('nodeName', '')), spec.get('nodeSelector', '')))
        # xls_input_cell_by_key(sheet, current_row, 'nodeName', spec.get('nodeName', ''))
        # xls_input_cell_by_key(sheet, current_row, 'role', hostname2role(spec.get('nodeName', '')))
        xls_input_cell_by_key(sheet, current_row, 'nodeSelector', dict2yaml(spec.get('nodeSelector', '')))
        print('  preemptionPolicy:{}, priorityClassName:{}'.format(spec.get('preemptionPolicy', ''), spec.get('priorityClassName', '')))
        xls_input_cell_by_key(sheet, current_row, 'preemptionPolicy', spec.get('preemptionPolicy', ''))
        xls_input_cell_by_key(sheet, current_row, 'priority', spec.get('priority', ''))
        xls_input_cell_by_key(sheet, current_row, 'priorityClassName', spec.get('priorityClassName', ''))
        xls_input_cell_by_key(sheet, current_row, 'restartPolicy', spec.get('restartPolicy', ''))
        xls_input_cell_by_key(sheet, current_row, 'schedulerName', spec.get('schedulerName', ''))
        xls_input_cell_by_key(sheet, current_row, 'serviceAccount', spec.get('serviceAccount', ''))
        xls_input_cell_by_key(sheet, current_row, 'serviceAccountName', spec.get('serviceAccountName', ''))
        print('  pod_securityContext:{}'.format(spec.get('securityContext', '')))
        xls_input_cell_by_key(sheet, current_row, 'pod_securityContext', dict2yaml(spec.get('securityContext', '')))
        print('  tolerations:{}'.format(spec.get('tolerations', '')))
        xls_input_cell_by_key(sheet, current_row, 'tolerations', dict2yaml(spec.get('tolerations', '')))
        xls_input_cell_by_key(sheet, current_row, 'terminationGracePeriodSeconds', spec.get('terminationGracePeriodSeconds', ''))
        print('  qosClass:{}'.format(status.get('qosClass', '')))
        xls_input_cell_by_key(sheet, current_row, 'qosClass', status.get('qosClass', ''))

        nctrs = 0
        row_pod_container_start = 0
        for ctr in spec.get('initContainers', list()):
            make_container_cells(sheet, current_row, ctr, True)
            if nctrs == 0:
                print('    ctr_name:{}, ctr_resources:{}, ctr_securityContext:{}'.format(
                    ctr.get('name', ''),
                    ctr.get('resources', ''),
                    ctr.get('securityContext', '')
                ))
                row_pod_container_start = current_row
            else:
                for col in range(header2column['ns'], header2column['qosClass'] + 1):
                    sheet.merge_cells(start_row=row_pod_container_start, end_row=current_row, start_column=col, end_column=col)

            nctrs += 1
            current_row = current_row + 1

        for ctr in spec.get('containers', list()):
            make_container_cells(sheet, current_row, ctr, False)
            if nctrs == 0:
                print('    ctr_name:{}, ctr_resources:{}, ctr_securityContext:{}'.format(
                    ctr.get('name', ''),
                    ctr.get('resources', ''),
                    ctr.get('securityContext', '')
                ))
                row_pod_container_start = current_row
            else:
                for col in range(header2column['ns'], header2column['qosClass'] + 1):
                    sheet.merge_cells(start_row=row_pod_container_start, end_row=current_row, start_column=col, end_column=col)

            nctrs += 1
            current_row = current_row + 1

    def get_cell_length(value):
        # print('*** get_cell_length(): value={}'.format(value))
        if str(value).find('\n'):
            sentences = str(value).split('\n')
            # print('*** sentences={}'.format(sentences))
            max_length = 0
            for s in sentences:
                if len(s) > max_length:
                    max_length = len(s)
            return max_length
        return len(str(value))

    for col in sheet.columns:
        max_length = 0
        colname = col[0].column_letter
        # if col[0].value == 'container_image':
        #     continue

        for cell in col:
            length = get_cell_length(cell.value)
            if length > max_length:
                   max_length = length

        target_width = 1
        if col[0].value == 'description':
            target_width = 60
        elif col[0].value == 'url':
            target_width = 30
        elif col[0].value == 'custom_resources':
            target_width = 30
        elif col[0].value == 'how_to_install':
            target_width = 30
        elif col[0].value == 'affinity':
            target_width = 30
        elif col[0].value == 'container_image':
            target_width = (len(str(col[0].value)) + 2) * 1.2
        else:
            target_width = (max_length + 2) * 1.2

        sheet.column_dimensions[colname].width = target_width

    book.save(args.output)

def argparse_debug(args):
    print('* args: {}'.format(args))
    print('* --online:', args.online)
    print('* --offline:', args.offline)
    print('* --description-yaml', args.description_yaml)
    print('* --output:', args.output)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--online', action='store_true', default=True)
    parser.add_argument('--offline', nargs='+', action='extend')
    parser.add_argument('--description-yaml', default='./description.yaml')
    parser.add_argument('--output', default='./newresult.xlsx')
    args = parser.parse_args()

    if args.offline:
        print('* running in offline mode...')
        args.online = False
    else:
        print('* running in online mode...')
        output = subprocess.run('oc whoami'.split(), capture_output=True)
        if output.returncode != 0:
            print('* oc command failed, exit')
            sys.exit(1)

    argparse_debug(args)

    main(args)
    sys.exit()
